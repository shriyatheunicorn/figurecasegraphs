# === PASS 1: Base workbook (Raw Data + Summary) ===

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

shutil.copy('/mnt/user-data/uploads/_Figure_Case_Study__Logistics_collect_data.xlsx', '/home/shriya/analysis.xlsx')
wb = openpyxl.load_workbook('/home/shriya/analysis.xlsx')
ws_raw = wb.active
ws_raw.title = "Raw Data"

last_row = ws_raw.max_row
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1E2761')

# Add implied cycle time column
ws_raw['F2'] = 'Implied Cycle Time (sec)'
ws_raw['F2'].font = header_font
ws_raw['F2'].fill = header_fill
ws_raw['F2'].alignment = Alignment(horizontal='center')
ws_raw.column_dimensions['F'].width = 22

for row in range(3, last_row + 1):
    ws_raw[f'F{row}'] = f'=D{row}/10'
    ws_raw[f'F{row}'].number_format = '0.000'

# Summary sheet
ws = wb.create_sheet("Summary Analysis")
ws['A1'] = 'LOGISTICS DATA COLLECTION \u2014 SUMMARY ANALYSIS'
ws['A1'].font = Font(bold=True, size=14, color='1E2761')
ws.merge_cells('A1:D1')

ws['A3'] = 'Overall Statistics'
ws['A3'].font = Font(bold=True, size=12, color='1E2761')

for i, h in enumerate(['Metric', 'Value']):
    c = ws.cell(row=4, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill

stats = [
    ('Total Episodes', f"=COUNTA('Raw Data'!D3:D{last_row})", '0'),
    ('Mean Episode Length (sec)', f"=AVERAGE('Raw Data'!D3:D{last_row})", '0.00'),
    ('Median Episode Length (sec)', f"=MEDIAN('Raw Data'!D3:D{last_row})", '0.00'),
    ('Std Dev (sec)', f"=STDEV('Raw Data'!D3:D{last_row})", '0.00'),
    ('Min Episode Length (sec)', f"=MIN('Raw Data'!D3:D{last_row})", '0.000'),
    ('Max Episode Length (sec)', f"=MAX('Raw Data'!D3:D{last_row})", '0.000'),
    ('Mean Implied Cycle Time (sec)', f"=AVERAGE('Raw Data'!D3:D{last_row})/10", '0.00'),
    ('Target Cycle Time (sec)', 3.5, '0.0'),
    ('% Under Target (35s ep. length)', f"=COUNTIF('Raw Data'!D3:D{last_row},\"<35\")/COUNTA('Raw Data'!D3:D{last_row})", '0.0%'),
    ('5th Percentile (sec)', f"=PERCENTILE('Raw Data'!D3:D{last_row},0.05)", '0.00'),
    ('95th Percentile (sec)', f"=PERCENTILE('Raw Data'!D3:D{last_row},0.95)", '0.00'),
    ('Outliers < 5s', f"=COUNTIF('Raw Data'!D3:D{last_row},\"<5\")", '0'),
    ('Outliers > 80s', f"=COUNTIF('Raw Data'!D3:D{last_row},\">80\")", '0'),
]

for i, (label, formula, fmt) in enumerate(stats):
    ws.cell(row=5+i, column=1, value=label)
    c = ws.cell(row=5+i, column=2)
    c.value = formula
    c.number_format = fmt

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 18

# Speed Mode section
ws['A20'] = 'Speed Mode Comparison'
ws['A20'].font = Font(bold=True, size=12, color='1E2761')

for i, h in enumerate(['Metric', 'Speed Episodes', 'Normal Episodes']):
    c = ws.cell(row=21, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill

ws['A22'] = 'Count'
ws['B22'] = f"=COUNTIF('Raw Data'!B3:B{last_row},\"*speed*\")"
ws['C22'] = f"=COUNTA('Raw Data'!D3:D{last_row})-B22"

ws['A23'] = 'Mean Episode Length (sec)'
ws['B23'] = f"=AVERAGEIF('Raw Data'!B3:B{last_row},\"*speed*\",'Raw Data'!D3:D{last_row})"
ws['C23'] = f"=(SUM('Raw Data'!D3:D{last_row})-SUMIF('Raw Data'!B3:B{last_row},\"*speed*\",'Raw Data'!D3:D{last_row}))/C22"
ws['B23'].number_format = '0.00'
ws['C23'].number_format = '0.00'

ws['A24'] = 'Implied Cycle Time (sec)'
ws['B24'] = '=B23/10'
ws['C24'] = '=C23/10'
ws['B24'].number_format = '0.00'
ws['C24'].number_format = '0.00'

ws.column_dimensions['C'].width = 20

# Highlight target row
ws['B12'].font = Font(bold=True, color='0D9488')

wb.save('/home/shriya/analysis.xlsx')
print("Spreadsheet saved!")

# === PASS 2: Extended sheets (Parsed Data, Operator, Config, Operator×Config, Daily) ===

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import re

# Load the existing workbook
wb = openpyxl.load_workbook('/home/shriya/analysis.xlsx')

# Parse the data with pandas for the new sheets
df = pd.read_excel('/mnt/user-data/uploads/_Figure_Case_Study__Logistics_collect_data.xlsx', skiprows=1)
df = df.drop(columns=['Unnamed: 0'])
df.columns = ['Dataset_name', 'Episode_ID', 'Episode_length_sec']
df['Episode_length_sec'] = pd.to_numeric(df['Episode_length_sec'], errors='coerce')
df = df.dropna(subset=['Dataset_name'])

def parse(name):
    info = {}
    info['date'] = name.split('_')[0]
    pilot_match = re.search(r'(Pilot\d+)', name)
    info['pilot'] = pilot_match.group(1) if pilot_match else None
    config_matches = re.findall(r'(c-\d+)', name)
    info['config'] = config_matches[-1] if config_matches else None
    info['quality'] = 'hq' if '_hq' in name else 'abc'
    info['no_bb'] = 'yes' if 'no_bb' in name.lower() else 'no'
    info['speed'] = 'yes' if 'speed' in name.lower() else 'no'
    return info

parsed = df['Dataset_name'].apply(parse).apply(pd.Series)
df = pd.concat([df, parsed], axis=1)

main = df[~df['Dataset_name'].str.contains('package|box', case=False)].copy()
main['implied_cycle'] = main['Episode_length_sec'] / 10
main['timestamp'] = pd.to_datetime(main['Episode_ID'].str[:14], format='%Y%m%d%H%M%S', errors='coerce')

# Styling helpers
header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
header_fill = PatternFill('solid', fgColor='1E2761')
alt_fill = PatternFill('solid', fgColor='F1F5F9')
accent_font = Font(bold=True, color='0D9488', name='Arial', size=10)
label_font = Font(name='Arial', size=10)
num_fmt_2 = '0.00'
num_fmt_0 = '#,##0'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style='thin', color='DEE2E6'),
    right=Side(style='thin', color='DEE2E6'),
    top=Side(style='thin', color='DEE2E6'),
    bottom=Side(style='thin', color='DEE2E6')
)

def style_header(ws, row, cols):
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        c.font = label_font
        c.border = thin_border
        if alt:
            c.fill = alt_fill

# =============================================
# SHEET: Parsed Data
# =============================================
ws_parsed = wb.create_sheet("Parsed Data")
ws_parsed['A1'] = 'PARSED EPISODE DATA (excluding early package/box pilot episodes)'
ws_parsed['A1'].font = Font(bold=True, size=12, color='1E2761', name='Arial')
ws_parsed.merge_cells('A1:J1')

headers = ['Dataset Name', 'Episode ID', 'Episode Length (sec)', 'Date', 'Config', 'Pilot', 'Quality', 'No BB', 'Speed Mode', 'Implied Cycle Time (sec)']
for i, h in enumerate(headers):
    ws_parsed.cell(row=3, column=i+1, value=h)
style_header(ws_parsed, 3, len(headers))

for idx, (_, row) in enumerate(main.iterrows()):
    r = idx + 4
    ws_parsed.cell(row=r, column=1, value=row['Dataset_name'])
    ws_parsed.cell(row=r, column=2, value=row['Episode_ID'])
    ws_parsed.cell(row=r, column=3, value=row['Episode_length_sec']).number_format = num_fmt_2
    ws_parsed.cell(row=r, column=4, value=row['date'])
    ws_parsed.cell(row=r, column=5, value=row['config'])
    ws_parsed.cell(row=r, column=6, value=row['pilot'])
    ws_parsed.cell(row=r, column=7, value=row['quality'])
    ws_parsed.cell(row=r, column=8, value=row['no_bb'])
    ws_parsed.cell(row=r, column=9, value=row['speed'])
    ws_parsed.cell(row=r, column=10, value=row['implied_cycle']).number_format = '0.000'
    style_data_row(ws_parsed, r, len(headers), alt=(idx % 2 == 1))

col_widths = [55, 35, 18, 12, 10, 10, 10, 8, 10, 22]
for i, w in enumerate(col_widths):
    ws_parsed.column_dimensions[get_column_letter(i+1)].width = w

last_parsed = len(main) + 3  # last data row

# =============================================
# SHEET: Operator Analysis
# =============================================
ws_op = wb.create_sheet("Operator Analysis")
ws_op['A1'] = 'TELEOPERATOR PERFORMANCE BREAKDOWN'
ws_op['A1'].font = Font(bold=True, size=12, color='1E2761', name='Arial')
ws_op.merge_cells('A1:H1')

op_headers = ['Pilot', 'Episodes', 'Mean Ep Length (sec)', 'Median Ep Length (sec)', 'Std Dev (sec)', 'Implied Cycle (sec)', '% Under Target', 'Active Dates']
for i, h in enumerate(op_headers):
    ws_op.cell(row=3, column=i+1, value=h)
style_header(ws_op, 3, len(op_headers))

pilot_stats = main.groupby('pilot').agg(
    count=('Episode_length_sec', 'count'),
    mean=('Episode_length_sec', 'mean'),
    median=('Episode_length_sec', 'median'),
    std=('Episode_length_sec', 'std'),
    cycle=('implied_cycle', 'mean'),
).reset_index().sort_values('cycle')

for idx, (_, row) in enumerate(pilot_stats.iterrows()):
    r = idx + 4
    sub = main[main['pilot'] == row['pilot']]
    pct_under = (sub['implied_cycle'] < 3.5).mean()
    dates = sorted(sub['date'].unique())
    date_str = ', '.join([f"{d[:4]}-{d[4:6]}-{d[6:]}" for d in dates])
    
    ws_op.cell(row=r, column=1, value=row['pilot'])
    ws_op.cell(row=r, column=2, value=int(row['count'])).number_format = num_fmt_0
    ws_op.cell(row=r, column=3, value=row['mean']).number_format = num_fmt_2
    ws_op.cell(row=r, column=4, value=row['median']).number_format = num_fmt_2
    ws_op.cell(row=r, column=5, value=row['std']).number_format = num_fmt_2
    ws_op.cell(row=r, column=6, value=row['cycle']).number_format = num_fmt_2
    ws_op.cell(row=r, column=7, value=pct_under).number_format = pct_fmt
    ws_op.cell(row=r, column=8, value=date_str)
    style_data_row(ws_op, r, len(op_headers), alt=(idx % 2 == 1))
    
    # Highlight if under target
    if row['cycle'] < 3.5:
        ws_op.cell(row=r, column=6).font = Font(bold=True, color='0D9488', name='Arial', size=10)

# Add formulas for totals/averages
r_total = len(pilot_stats) + 4
ws_op.cell(row=r_total, column=1, value='TOTAL / AVERAGE').font = Font(bold=True, name='Arial', size=10)
ws_op.cell(row=r_total, column=2, value=f'=SUM(B4:B{r_total-1})').number_format = num_fmt_0
ws_op.cell(row=r_total, column=3, value=f'=SUMPRODUCT(B4:B{r_total-1},C4:C{r_total-1})/B{r_total}').number_format = num_fmt_2
ws_op.cell(row=r_total, column=6, value=f'=SUMPRODUCT(B4:B{r_total-1},F4:F{r_total-1})/B{r_total}').number_format = num_fmt_2
for col in range(1, len(op_headers)+1):
    ws_op.cell(row=r_total, column=col).border = Border(top=Side(style='medium', color='1E2761'), bottom=Side(style='medium', color='1E2761'))

op_col_widths = [10, 10, 20, 20, 14, 18, 14, 55]
for i, w in enumerate(op_col_widths):
    ws_op.column_dimensions[get_column_letter(i+1)].width = w

# =============================================
# SHEET: Config Analysis
# =============================================
ws_cfg = wb.create_sheet("Config Analysis")
ws_cfg['A1'] = 'ROBOT CONFIGURATION PERFORMANCE BREAKDOWN'
ws_cfg['A1'].font = Font(bold=True, size=12, color='1E2761', name='Arial')
ws_cfg.merge_cells('A1:H1')

cfg_headers = ['Config', 'Episodes', 'Mean Ep Length (sec)', 'Median Ep Length (sec)', 'Std Dev (sec)', 'Implied Cycle (sec)', '% Under Target', 'Active Dates']
for i, h in enumerate(cfg_headers):
    ws_cfg.cell(row=3, column=i+1, value=h)
style_header(ws_cfg, 3, len(cfg_headers))

config_stats = main.groupby('config').agg(
    count=('Episode_length_sec', 'count'),
    mean=('Episode_length_sec', 'mean'),
    median=('Episode_length_sec', 'median'),
    std=('Episode_length_sec', 'std'),
    cycle=('implied_cycle', 'mean'),
).reset_index().sort_values('cycle')

for idx, (_, row) in enumerate(config_stats.iterrows()):
    r = idx + 4
    sub = main[main['config'] == row['config']]
    pct_under = (sub['implied_cycle'] < 3.5).mean()
    dates = sorted(sub['date'].unique())
    date_str = ', '.join([f"{d[:4]}-{d[4:6]}-{d[6:]}" for d in dates])
    
    ws_cfg.cell(row=r, column=1, value=row['config'])
    ws_cfg.cell(row=r, column=2, value=int(row['count'])).number_format = num_fmt_0
    ws_cfg.cell(row=r, column=3, value=row['mean']).number_format = num_fmt_2
    ws_cfg.cell(row=r, column=4, value=row['median']).number_format = num_fmt_2
    ws_cfg.cell(row=r, column=5, value=row['std']).number_format = num_fmt_2
    ws_cfg.cell(row=r, column=6, value=row['cycle']).number_format = num_fmt_2
    ws_cfg.cell(row=r, column=7, value=pct_under).number_format = pct_fmt
    ws_cfg.cell(row=r, column=8, value=date_str)
    style_data_row(ws_cfg, r, len(cfg_headers), alt=(idx % 2 == 1))
    
    if row['cycle'] < 3.5:
        ws_cfg.cell(row=r, column=6).font = Font(bold=True, color='0D9488', name='Arial', size=10)

r_total = len(config_stats) + 4
ws_cfg.cell(row=r_total, column=1, value='TOTAL / AVERAGE').font = Font(bold=True, name='Arial', size=10)
ws_cfg.cell(row=r_total, column=2, value=f'=SUM(B4:B{r_total-1})').number_format = num_fmt_0
ws_cfg.cell(row=r_total, column=3, value=f'=SUMPRODUCT(B4:B{r_total-1},C4:C{r_total-1})/B{r_total}').number_format = num_fmt_2
ws_cfg.cell(row=r_total, column=6, value=f'=SUMPRODUCT(B4:B{r_total-1},F4:F{r_total-1})/B{r_total}').number_format = num_fmt_2
for col in range(1, len(cfg_headers)+1):
    ws_cfg.cell(row=r_total, column=col).border = Border(top=Side(style='medium', color='1E2761'), bottom=Side(style='medium', color='1E2761'))

cfg_col_widths = [10, 10, 20, 20, 14, 18, 14, 45]
for i, w in enumerate(cfg_col_widths):
    ws_cfg.column_dimensions[get_column_letter(i+1)].width = w

# =============================================
# SHEET: Operator x Config
# =============================================
ws_ox = wb.create_sheet("Operator x Config")
ws_ox['A1'] = 'OPERATOR \u00d7 CONFIGURATION CROSS-TABULATION'
ws_ox['A1'].font = Font(bold=True, size=12, color='1E2761', name='Arial')
ws_ox.merge_cells('A1:H1')

ox_headers = ['Pilot', 'Config', 'Episodes', 'Mean Ep Length (sec)', 'Implied Cycle (sec)', 'First Timestamp', 'Last Timestamp']
for i, h in enumerate(ox_headers):
    ws_ox.cell(row=3, column=i+1, value=h)
style_header(ws_ox, 3, len(ox_headers))

pairs = main.groupby(['pilot', 'config']).agg(
    count=('Episode_length_sec', 'count'),
    mean=('Episode_length_sec', 'mean'),
    cycle=('implied_cycle', 'mean'),
    first_ts=('timestamp', 'min'),
    last_ts=('timestamp', 'max')
).reset_index().sort_values(['pilot', 'cycle'])

for idx, (_, row) in enumerate(pairs.iterrows()):
    r = idx + 4
    ws_ox.cell(row=r, column=1, value=row['pilot'])
    ws_ox.cell(row=r, column=2, value=row['config'])
    ws_ox.cell(row=r, column=3, value=int(row['count'])).number_format = num_fmt_0
    ws_ox.cell(row=r, column=4, value=row['mean']).number_format = num_fmt_2
    ws_ox.cell(row=r, column=5, value=row['cycle']).number_format = num_fmt_2
    ws_ox.cell(row=r, column=6, value=row['first_ts'].strftime('%Y-%m-%d %H:%M:%S') if pd.notna(row['first_ts']) else '')
    ws_ox.cell(row=r, column=7, value=row['last_ts'].strftime('%Y-%m-%d %H:%M:%S') if pd.notna(row['last_ts']) else '')
    style_data_row(ws_ox, r, len(ox_headers), alt=(idx % 2 == 1))
    
    if row['cycle'] < 3.5:
        ws_ox.cell(row=r, column=5).font = Font(bold=True, color='0D9488', name='Arial', size=10)

ox_col_widths = [10, 10, 10, 20, 18, 22, 22]
for i, w in enumerate(ox_col_widths):
    ws_ox.column_dimensions[get_column_letter(i+1)].width = w

# =============================================
# SHEET: Daily Trends
# =============================================
ws_daily = wb.create_sheet("Daily Trends")
ws_daily['A1'] = 'DAILY PERFORMANCE TRENDS'
ws_daily['A1'].font = Font(bold=True, size=12, color='1E2761', name='Arial')
ws_daily.merge_cells('A1:G1')

daily_headers = ['Date', 'Episodes', 'Mean Ep Length (sec)', 'Median Ep Length (sec)', 'Std Dev (sec)', 'Implied Cycle (sec)', '% Under Target']
for i, h in enumerate(daily_headers):
    ws_daily.cell(row=3, column=i+1, value=h)
style_header(ws_daily, 3, len(daily_headers))

date_stats = main.groupby('date').agg(
    count=('Episode_length_sec', 'count'),
    mean=('Episode_length_sec', 'mean'),
    median=('Episode_length_sec', 'median'),
    std=('Episode_length_sec', 'std'),
    cycle=('implied_cycle', 'mean'),
).reset_index().sort_values('date')

for idx, (_, row) in enumerate(date_stats.iterrows()):
    r = idx + 4
    d = row['date']
    date_formatted = f"{d[:4]}-{d[4:6]}-{d[6:]}"
    sub = main[main['date'] == row['date']]
    pct_under = (sub['implied_cycle'] < 3.5).mean()
    
    ws_daily.cell(row=r, column=1, value=date_formatted)
    ws_daily.cell(row=r, column=2, value=int(row['count'])).number_format = num_fmt_0
    ws_daily.cell(row=r, column=3, value=row['mean']).number_format = num_fmt_2
    ws_daily.cell(row=r, column=4, value=row['median']).number_format = num_fmt_2
    ws_daily.cell(row=r, column=5, value=row['std']).number_format = num_fmt_2
    ws_daily.cell(row=r, column=6, value=row['cycle']).number_format = num_fmt_2
    ws_daily.cell(row=r, column=7, value=pct_under).number_format = pct_fmt
    style_data_row(ws_daily, r, len(daily_headers), alt=(idx % 2 == 1))
    
    if row['cycle'] < 3.5:
        ws_daily.cell(row=r, column=6).font = Font(bold=True, color='0D9488', name='Arial', size=10)

daily_col_widths = [14, 10, 20, 20, 14, 18, 14]
for i, w in enumerate(daily_col_widths):
    ws_daily.column_dimensions[get_column_letter(i+1)].width = w

# Save
wb.save('/home/shriya/analysis.xlsx')
print("Done \u2014 4 new sheets added!")